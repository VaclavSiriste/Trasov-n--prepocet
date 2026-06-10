#!/usr/bin/env python3
"""Import období, objednáno, denní rozpis montérů a ručního zápisu z Excelu."""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "server"))
sys.path.insert(0, str(ROOT.parent / ".pylibs"))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor


DEFAULT_XLSX = Path.home() / "Downloads" / "Kopie souboru Trasování (1).xlsx"
MONTH_SHEETS = {
    (2026, 6): "Červen 2026",
}


def parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return date.fromisoformat(str(val)[:10])
    except ValueError:
        return None


def target_flag(val) -> int:
    if val in (1, 1.0, True):
        return 1
    return 0


def import_prehled_obdobi(cur, ws, update_dates: bool = False) -> int:
    """Import ručních sloupců N (objednáno) a O (posunout výrobu). Data od–do řídí číselník skupin v DB."""
    updated = 0
    for r in range(2, 20):
        lok = ws.cell(r, 5).value
        if not lok or lok not in {
            "MSK", "PR/ST", "BR", "PCE/KH", "PL", "Ústí", "Libr",
            "Zlín", "Olomouc", "Vysočina", "České Budějovice",
        }:
            continue
        skupina = ws.cell(r, 3).value
        objednano = ws.cell(r, 14).value
        posunout = ws.cell(r, 15).value or "NE"
        if update_dates:
            od = parse_date(ws.cell(r, 1).value)
            do = parse_date(ws.cell(r, 2).value)
            cur.execute(
                """UPDATE prehled_obdobi SET od=%s, "do"=%s, skupina=%s,
                   objednano_ks=%s, posunout_vyrobu=%s WHERE lokalita=%s""",
                (od, do, skupina, float(objednano or 0), str(posunout), lok),
            )
        else:
            cur.execute(
                """UPDATE prehled_obdobi SET skupina=COALESCE(%s, skupina),
                   objednano_ks=%s, posunout_vyrobu=%s WHERE lokalita=%s""",
                (skupina, float(objednano or 0), str(posunout), lok),
            )
        updated += cur.rowcount
    return updated


def import_daily_roster(cur, ws, mesic_key: str) -> int:
    names: list[tuple[int, str]] = []
    for c in range(16, 260, 4):
        name = ws.cell(2, c).value
        if name and str(name).strip():
            names.append((c, str(name).strip()))

    cur.execute("DELETE FROM mesicni_rozpis_den WHERE mesic_key = %s", (mesic_key,))
    count = 0
    for r in range(4, 36):
        day = parse_date(ws.cell(r, 1).value)
        if not day:
            continue
        for base, name in names:
            target = ws.cell(r, base + 1).value
            kraj = ws.cell(r, base + 2).value
            if kraj is None and target is None:
                continue
            cur.execute(
                """INSERT INTO mesicni_rozpis_den
                   (mesic_key, col_index, jmeno, datum, target_flag, destination_region)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (mesic_key, base, name, day, target_flag(target), str(kraj or "MSK")),
            )
            count += 1
    return count


def import_zapis_den(cur, ws, mesic_key: str) -> int:
    cur.execute("DELETE FROM mesicni_zapis_den WHERE mesic_key = %s", (mesic_key,))
    count = 0
    for r in range(4, 36):
        day = parse_date(ws.cell(r, 1).value)
        if not day:
            continue
        collected = ws.cell(r, 7).value
        reason = ws.cell(r, 5).value
        if collected in (None, "") and not reason:
            continue
        try:
            coll = float(collected or 0)
        except (TypeError, ValueError):
            coll = 0
        cur.execute(
            """INSERT INTO mesicni_zapis_den (mesic_key, datum, collected, reason)
               VALUES (%s, %s, %s, %s)""",
            (mesic_key, day, coll, str(reason or "")),
        )
        count += 1
    return count


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--mesic-key", default="2026-06")
    parser.add_argument("--dsn", default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--update-dates", action="store_true", help="Přepsat i sloupce Od/Do z Excelu")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Excel nenalezen: {args.xlsx}")

    rok, mesic = map(int, args.mesic_key.split("-"))
    sheet = MONTH_SHEETS.get((rok, mesic))
    if not sheet:
        raise SystemExit(f"Neznámý měsíc: {args.mesic_key}")

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws_prehled = wb["Přehled MONTÁŽE"]
    ws_mesic = wb[sheet]

    conn = psycopg2.connect(args.dsn, cursor_factory=RealDictCursor)
    cur = conn.cursor()

    for sql_file in ("migrate_mesicni_rozpis_den.sql", "migrate_mesicni_zapis_den.sql"):
        path = ROOT / "sql" / sql_file
        if path.exists():
            cur.execute(path.read_text())

    obdobi = import_prehled_obdobi(cur, ws_prehled, update_dates=args.update_dates)
    roster = import_daily_roster(cur, ws_mesic, args.mesic_key)
    zapis = import_zapis_den(cur, ws_mesic, args.mesic_key)
    conn.commit()
    wb.close()
    conn.close()

    print(f"prehled_obdobi aktualizováno: {obdobi}")
    print(f"mesicni_rozpis_den: {roster} záznamů")
    print(f"mesicni_zapis_den: {zapis} záznamů")


if __name__ == "__main__":
    main()
